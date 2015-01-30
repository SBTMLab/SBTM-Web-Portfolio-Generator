 # -*- coding: utf-8 -*-

from jinja2 import Environment, FileSystemLoader
import os, shutil
from openpyxl import load_workbook
from resizer import resize


wb = load_workbook(filename = 'sbtmwebport.xlsx')

env = Environment(loader=FileSystemLoader(u'./templates'))

if os.path.isdir(u"./result"):
	shutil.rmtree(u"./result")
shutil.copytree(u"./static","./result")


features = []
indexsheet =  wb['index.html']

featindex = 6
filename = indexsheet['C'+str(featindex)].value

while filename :
	features.append(filename)
	featindex += 1
	filename = indexsheet['C'+str(featindex)].value

index = env.get_template('index.html')

indexhtml = index.render(features=features).encode('utf-8')



pieces = []

projectsheet =  wb['projects.html']

projectindex = 5
filename = projectsheet['B'+str(projectindex)].value
while filename :
	tag = projectsheet['C'+str(projectindex)].value
	maker = projectsheet['D'+str(projectindex)].value
	year = projectsheet['E'+str(projectindex)].value
	objective = projectsheet['F'+str(projectindex)].value
	link = projectsheet['G'+str(projectindex)].value
	headtext = projectsheet['H'+str(projectindex)].value
	maintext = projectsheet['I'+str(projectindex)].value

	imgs =[]
	col = 10
	img = dict(origin =projectsheet.cell(row = projectindex, column = col).value)
	
	while img['origin'] :
		img["thumb"] = resize(img["origin"],"./projectimage/","./result/",940,529)
		imgs.append(img)
		col += 1
		img = dict(origin =projectsheet.cell(row = projectindex, column = col).value)

	page = tag + ".html"
	p = dict (name = filename, tag = tag, maker = maker, year = year, objective = objective, link = link, headtext = headtext, maintext = maintext, imgs = imgs, page = page)

	pieces.append(p)
	projectindex += 1
	filename = projectsheet['B'+str(projectindex)].value

projects = env.get_template('projects.html')

projectshtml = projects.render(pieces = pieces ).encode('utf-8')


project = env.get_template('project-details.html')




for pc in pieces :
	ph = project.render(piece = pc , otherprojects = pieces).encode('utf-8')
	f = open(u"./result/"+pc['page'], 'w') 
	f.write(ph)
	f.close()

contact = env.get_template('contact.html')
contacthtml = contact.render().encode('utf-8')


f = open(u"./result/index.html", 'w') 
f.write(indexhtml)
f.close()
f = open(u"./result/projects.html", 'w') 
f.write(projectshtml)
f.close()

f = open(u"./result/contact.html", 'w') 
f.write(contacthtml)
f.close()

